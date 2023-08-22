from openpyxl import load_workbook

from tests.conftest import path_to_file_res

xlsxfile = path_to_file_res('file_example_XLSX_50.xlsx')
# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx_file():
    workbook = load_workbook(xlsxfile)
    sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == 'Mara'
